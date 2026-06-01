# tests/test_metricas_service.py
from datetime import datetime, timedelta, timezone

from app.core.security import get_password_hash
from app.models.user import RoleEnum, User
from app.schemas.incidente import (
    AccionCorrectivaCreate,
    AccionCorrectivaUpdate,
    IncidenteCreate,
    InvestigacionCreate,
    LesionCreate,
)
from app.services import incidente_service, metricas_service

# ── Helpers ─────────────────────────────────────────────────────────


def make_empleado(db, empresa):
    import secrets

    user = User(
        nombre="Empleado Test",
        email=f"emp_{secrets.token_hex(4)}@test.com",
        password_hash=get_password_hash("Password1!"),
        role=RoleEnum.empleado,
        empresa_id=empresa.id,
        activo=True,
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    return user


def make_incidente(db, empresa, usuario_sst, tipo="accidente", **kwargs):
    datos = IncidenteCreate(
        tipo=tipo,
        severidad="leve",
        fecha=datetime.now(timezone.utc),
        lugar="Planta A",
        descripcion="Descripción del incidente",
        **kwargs,
    )
    return incidente_service.create_incidente(db, datos, empresa.id, usuario_sst.id)


def make_accion(db, empresa, incidente, usuario_sst, fecha_limite=None):
    if fecha_limite is None:
        fecha_limite = datetime.now(timezone.utc) + timedelta(days=30)
    datos = AccionCorrectivaCreate(
        descripcion="Acción correctiva",
        fecha_limite=fecha_limite,
        responsable_id=usuario_sst.id,
    )
    return incidente_service.create_accion_correctiva(
        db, incidente.id, empresa.id, datos
    )


# ── get_kpis ─────────────────────────────────────────────────────────


def test_kpis_empresa_sin_datos(db, empresa):
    resultado = metricas_service.get_kpis(db, empresa.id)
    assert resultado["total_trabajadores"] == 0
    assert resultado["total_accidentes"] == 0
    assert resultado["dias_perdidos"] == 0
    assert resultado["tasa_accidentalidad"] == 0
    assert resultado["indice_frecuencia"] == 0
    assert resultado["indice_severidad"] == 0


def test_kpis_con_empleados(db, empresa, usuario_sst):
    make_empleado(db, empresa)
    make_empleado(db, empresa)
    resultado = metricas_service.get_kpis(db, empresa.id)
    assert resultado["total_trabajadores"] == 2


def test_kpis_con_accidentes(db, empresa, usuario_sst):
    make_empleado(db, empresa)
    make_incidente(db, empresa, usuario_sst, tipo="accidente")
    resultado = metricas_service.get_kpis(db, empresa.id)
    assert resultado["total_accidentes"] == 1
    assert resultado["tasa_accidentalidad"] > 0
    assert resultado["indice_frecuencia"] > 0


def test_kpis_incidente_no_accidente_no_cuenta(db, empresa, usuario_sst):
    make_empleado(db, empresa)
    make_incidente(db, empresa, usuario_sst, tipo="incidente")
    resultado = metricas_service.get_kpis(db, empresa.id)
    assert resultado["total_accidentes"] == 0


def test_kpis_con_dias_perdidos(db, empresa, usuario_sst):
    make_empleado(db, empresa)
    datos = IncidenteCreate(
        tipo="accidente",
        severidad="grave",
        fecha=datetime.now(timezone.utc),
        lugar="Bodega",
        descripcion="Accidente grave",
        lesion=LesionCreate(
            tipo_lesion="fractura", parte_afectada="brazo", incapacidad_dias=10
        ),
    )
    incidente_service.create_incidente(db, datos, empresa.id, usuario_sst.id)
    resultado = metricas_service.get_kpis(db, empresa.id)
    assert resultado["dias_perdidos"] == 10
    assert resultado["indice_severidad"] > 0


# ── get_dashboard_gerencia ───────────────────────────────────────────


def test_dashboard_empresa_vacia(db, empresa):
    resultado = metricas_service.get_dashboard_gerencia(db, empresa.id)
    assert resultado["cumplimiento_sgsst"] == 100
    assert resultado["incidentes_activos"] == 0
    assert resultado["total_capacitaciones"] == 0
    assert resultado["acciones_vencidas"] == 0
    assert "kpis" in resultado


def test_dashboard_con_incidente_activo(db, empresa, usuario_sst):
    inc = make_incidente(db, empresa, usuario_sst)
    incidente_service.update_estado_incidente(db, inc.id, empresa.id, "abierto")
    resultado = metricas_service.get_dashboard_gerencia(db, empresa.id)
    assert resultado["incidentes_activos"] >= 1


def test_dashboard_cumplimiento_con_acciones(db, empresa, usuario_sst):
    inc = make_incidente(db, empresa, usuario_sst)
    accion = make_accion(db, empresa, inc, usuario_sst)
    incidente_service.update_accion_correctiva(
        db,
        accion.id,
        empresa.id,
        AccionCorrectivaUpdate(estado="completada", evidencia="Evidencia"),
    )
    resultado = metricas_service.get_dashboard_gerencia(db, empresa.id)
    assert resultado["cumplimiento_sgsst"] == 100


def test_dashboard_acciones_vencidas(db, empresa, usuario_sst):
    inc = make_incidente(db, empresa, usuario_sst)
    fecha_pasada = datetime.now(timezone.utc) - timedelta(days=5)
    make_accion(db, empresa, inc, usuario_sst, fecha_limite=fecha_pasada)
    resultado = metricas_service.get_dashboard_gerencia(db, empresa.id)
    assert resultado["acciones_vencidas"] >= 1


def test_dashboard_incidentes_ultimo_mes(db, empresa, usuario_sst):
    make_incidente(db, empresa, usuario_sst)
    resultado = metricas_service.get_dashboard_gerencia(db, empresa.id)
    assert resultado["incidentes_ultimo_mes"] >= 1


# ── get_alertas ──────────────────────────────────────────────────────


def test_alertas_empresa_sin_problemas(db, empresa):
    resultado = metricas_service.get_alertas(db, empresa.id)
    assert resultado["total_alertas"] == 0
    assert resultado["alertas"] == []


def test_alerta_incidente_abierto_sin_investigacion(db, empresa, usuario_sst):
    inc = make_incidente(db, empresa, usuario_sst)
    incidente_service.update_estado_incidente(db, inc.id, empresa.id, "abierto")
    resultado = metricas_service.get_alertas(db, empresa.id)
    tipos = [a["tipo"] for a in resultado["alertas"]]
    assert "incidente_sin_investigacion" in tipos
    alerta = next(
        a for a in resultado["alertas"] if a["tipo"] == "incidente_sin_investigacion"
    )
    assert alerta["nivel"] == "critico"


def test_alerta_accion_vencida(db, empresa, usuario_sst):
    inc = make_incidente(db, empresa, usuario_sst)
    fecha_pasada = datetime.now(timezone.utc) - timedelta(days=3)
    make_accion(db, empresa, inc, usuario_sst, fecha_limite=fecha_pasada)
    resultado = metricas_service.get_alertas(db, empresa.id)
    tipos = [a["tipo"] for a in resultado["alertas"]]
    assert "accion_correctiva_vencida" in tipos
    alerta = next(
        a for a in resultado["alertas"] if a["tipo"] == "accion_correctiva_vencida"
    )
    assert alerta["nivel"] == "critico"


def test_alerta_accion_proxima(db, empresa, usuario_sst):
    inc = make_incidente(db, empresa, usuario_sst)
    fecha_proxima = datetime.now(timezone.utc) + timedelta(days=3)
    make_accion(db, empresa, inc, usuario_sst, fecha_limite=fecha_proxima)
    resultado = metricas_service.get_alertas(db, empresa.id)
    tipos = [a["tipo"] for a in resultado["alertas"]]
    assert "accion_correctiva_proxima" in tipos
    alerta = next(
        a for a in resultado["alertas"] if a["tipo"] == "accion_correctiva_proxima"
    )
    assert alerta["nivel"] == "medio"


def test_sin_alerta_si_incidente_tiene_investigacion(db, empresa, usuario_sst):
    inc = make_incidente(db, empresa, usuario_sst)
    incidente_service.update_estado_incidente(db, inc.id, empresa.id, "abierto")
    incidente_service.create_investigacion(
        db, inc.id, empresa.id, InvestigacionCreate(causas_inmediatas="Piso mojado")
    )
    resultado = metricas_service.get_alertas(db, empresa.id)
    tipos = [a["tipo"] for a in resultado["alertas"]]
    assert "incidente_sin_investigacion" not in tipos


def test_sin_alerta_si_accion_completada(db, empresa, usuario_sst):
    inc = make_incidente(db, empresa, usuario_sst)
    fecha_pasada = datetime.now(timezone.utc) - timedelta(days=3)
    accion = make_accion(db, empresa, inc, usuario_sst, fecha_limite=fecha_pasada)
    incidente_service.update_accion_correctiva(
        db,
        accion.id,
        empresa.id,
        AccionCorrectivaUpdate(estado="completada", evidencia="Evidencia"),
    )
    resultado = metricas_service.get_alertas(db, empresa.id)
    tipos = [a["tipo"] for a in resultado["alertas"]]
    assert "accion_correctiva_vencida" not in tipos


# ── generar_reporte_pdf ──────────────────────────────────────────────


def test_reporte_pdf_retorna_bytesio(db, empresa):
    resultado = metricas_service.generar_reporte_pdf(db, empresa.id, "mensual")
    contenido = resultado.read()
    assert len(contenido) > 0
    assert contenido[:4] == b"%PDF"


def test_reporte_pdf_periodos(db, empresa):
    for periodo in ["mensual", "trimestral", "anual"]:
        resultado = metricas_service.generar_reporte_pdf(db, empresa.id, periodo)
        assert resultado.read()[:4] == b"%PDF"


# ── generar_reporte_excel ────────────────────────────────────────────


def test_reporte_excel_retorna_bytesio(db, empresa):
    import openpyxl

    resultado = metricas_service.generar_reporte_excel(db, empresa.id, "mensual")
    wb = openpyxl.load_workbook(resultado)
    assert "Reporte PISST" in wb.sheetnames


def test_reporte_excel_contiene_kpis(db, empresa, usuario_sst):
    import openpyxl

    make_empleado(db, empresa)
    make_incidente(db, empresa, usuario_sst)
    resultado = metricas_service.generar_reporte_excel(db, empresa.id, "trimestral")
    wb = openpyxl.load_workbook(resultado)
    ws = wb["Reporte PISST"]
    valores = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
    assert any("KPI" in str(v) for v in valores if v)
