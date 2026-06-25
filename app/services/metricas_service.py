# app/services/metricas_service.py
from datetime import date, datetime, timedelta, timezone
from typing import Optional
from uuid import UUID

from sqlalchemy import func
from sqlalchemy.orm import Session

from app.models.accion_correctiva import AccionCorrectiva, EstadoAccionEnum
from app.models.capacitacion import Capacitacion
from app.models.incidente import EstadoIncidenteEnum, Incidente, TipoIncidenteEnum
from app.models.user import RoleEnum, User


def get_kpis(db: Session, empresa_id: UUID):
    total_trabajadores = (
        db.query(User)
        .filter(
            User.empresa_id == empresa_id,
            User.activo == True,
            User.role == RoleEnum.empleado,
        )
        .count()
    )

    inicio_anio = datetime(datetime.now(timezone.utc).replace(tzinfo=None).year, 1, 1)
    total_accidentes = (
        db.query(Incidente)
        .filter(
            Incidente.empresa_id == empresa_id,
            Incidente.tipo == TipoIncidenteEnum.accidente,
            Incidente.fecha >= inicio_anio,
        )
        .count()
    )

    from app.models.lesion import Lesion

    dias_perdidos_result = (
        db.query(func.sum(Lesion.incapacidad_dias))
        .join(Incidente, Lesion.incidente_id == Incidente.id)
        .filter(Incidente.empresa_id == empresa_id, Incidente.fecha >= inicio_anio)
        .scalar()
    )
    dias_perdidos = dias_perdidos_result or 0

    dias_transcurridos = (
        datetime.now(timezone.utc).replace(tzinfo=None) - inicio_anio
    ).days

    # ✅ Fix Bug #5 — División por cero el 1 de enero
    horas_trabajadas = total_trabajadores * 8 * dias_transcurridos
    horas_trabajadas = horas_trabajadas if horas_trabajadas > 0 else 1

    tasa_accidentalidad = (
        round((total_accidentes / total_trabajadores) * 100, 2)
        if total_trabajadores > 0
        else 0
    )
    indice_frecuencia = (
        round((total_accidentes / horas_trabajadas) * 1000000, 2)
        if horas_trabajadas > 0
        else 0
    )
    indice_severidad = (
        round((dias_perdidos / horas_trabajadas) * 1000000, 2)
        if horas_trabajadas > 0
        else 0
    )

    return {
        "total_trabajadores": total_trabajadores,
        "total_accidentes": total_accidentes,
        "dias_perdidos": dias_perdidos,
        "tasa_accidentalidad": tasa_accidentalidad,
        "indice_frecuencia": indice_frecuencia,
        "indice_severidad": indice_severidad,
    }


def get_dashboard_gerencia(
    db: Session,
    empresa_id: UUID,
    fecha_desde: Optional[date] = None,
    fecha_hasta: Optional[date] = None,
):
    kpis = get_kpis(db, empresa_id)

    incidentes_activos = (
        db.query(Incidente)
        .filter(
            Incidente.empresa_id == empresa_id,
            Incidente.estado != EstadoIncidenteEnum.cerrado,
        )
        .count()
    )

    if fecha_desde and fecha_hasta:
        desde_dt = datetime.combine(fecha_desde, datetime.min.time())
        hasta_dt = datetime.combine(fecha_hasta, datetime.max.time())
    else:
        hasta_dt = datetime.now(timezone.utc).replace(tzinfo=None)
        desde_dt = hasta_dt - timedelta(days=30)

    incidentes_ultimo_mes = (
        db.query(Incidente)
        .filter(
            Incidente.empresa_id == empresa_id,
            Incidente.fecha_creacion >= desde_dt,
            Incidente.fecha_creacion <= hasta_dt,
        )
        .count()
    )

    cap_query = db.query(Capacitacion).filter(
        Capacitacion.empresa_id == empresa_id, Capacitacion.activo == True
    )
    if fecha_desde and fecha_hasta:
        cap_query = cap_query.filter(
            Capacitacion.fecha_creacion >= desde_dt,
            Capacitacion.fecha_creacion <= hasta_dt,
        )
    total_capacitaciones = cap_query.count()

    acciones_base = db.query(AccionCorrectiva).join(
        Incidente, AccionCorrectiva.incidente_id == Incidente.id
    )
    if fecha_desde and fecha_hasta:
        acciones_base = acciones_base.filter(
            AccionCorrectiva.fecha_creacion >= desde_dt,
            AccionCorrectiva.fecha_creacion <= hasta_dt,
        )

    acciones_vencidas = acciones_base.filter(
        Incidente.empresa_id == empresa_id,
        AccionCorrectiva.estado != EstadoAccionEnum.completada,
        AccionCorrectiva.fecha_limite < datetime.now(timezone.utc).replace(tzinfo=None),
    ).count()

    total_acciones = acciones_base.filter(Incidente.empresa_id == empresa_id).count()

    acciones_completadas = acciones_base.filter(
        Incidente.empresa_id == empresa_id,
        AccionCorrectiva.estado == EstadoAccionEnum.completada,
    ).count()

    cumplimiento = (
        round((acciones_completadas / total_acciones) * 100)
        if total_acciones > 0
        else 100
    )

    return {
        "cumplimiento_sgsst": cumplimiento,
        "incidentes_activos": incidentes_activos,
        "incidentes_ultimo_mes": incidentes_ultimo_mes,
        "total_capacitaciones": total_capacitaciones,
        "acciones_vencidas": acciones_vencidas,
        "kpis": kpis,
    }


def get_alertas(db: Session, empresa_id: UUID):
    alertas = []

    sin_investigacion = (
        db.query(Incidente)
        .filter(
            Incidente.empresa_id == empresa_id,
            Incidente.estado == EstadoIncidenteEnum.abierto,
            ~Incidente.investigacion.has(),
        )
        .count()
    )

    if sin_investigacion > 0:
        alertas.append(
            {
                "tipo": "incidente_sin_investigacion",
                "nivel": "critico",
                "mensaje": f"{sin_investigacion} incidente(s) abiertos sin investigación documentada",
                "url_destino": "/incidentes?estado=abierto",
            }
        )

    ahora = datetime.now(timezone.utc).replace(tzinfo=None)

    _query_vencidas = (
        db.query(AccionCorrectiva)
        .join(Incidente, AccionCorrectiva.incidente_id == Incidente.id)
        .filter(
            Incidente.empresa_id == empresa_id,
            AccionCorrectiva.estado != EstadoAccionEnum.completada,
            AccionCorrectiva.fecha_limite < ahora,
        )
    )
    acciones_vencidas = _query_vencidas.count()

    if acciones_vencidas > 0:
        primera = _query_vencidas.first()
        alertas.append(
            {
                "tipo": "accion_correctiva_vencida",
                "nivel": "critico",
                "mensaje": f"{acciones_vencidas} acción(es) correctiva(s) vencida(s)",
                "url_destino": f"/incidentes?reporte={primera.incidente_id}",
            }
        )

    proxima_semana = ahora + timedelta(days=7)
    _query_proximas = (
        db.query(AccionCorrectiva)
        .join(Incidente, AccionCorrectiva.incidente_id == Incidente.id)
        .filter(
            Incidente.empresa_id == empresa_id,
            AccionCorrectiva.estado != EstadoAccionEnum.completada,
            AccionCorrectiva.fecha_limite >= ahora,
            AccionCorrectiva.fecha_limite <= proxima_semana,
        )
    )
    acciones_proximas = _query_proximas.count()

    if acciones_proximas > 0:
        primera_proxima = _query_proximas.first()
        alertas.append(
            {
                "tipo": "accion_correctiva_proxima",
                "nivel": "medio",
                "mensaje": f"{acciones_proximas} acción(es) correctiva(s) vence(n) en los próximos 7 días",
                "url_destino": f"/incidentes?reporte={primera_proxima.incidente_id}",
            }
        )

    return {"total_alertas": len(alertas), "alertas": alertas}


# ── Reportes ejecutivos ───────────────────────────────────────────


def generar_reporte_pdf(db: Session, empresa_id: UUID, periodo: str):
    from io import BytesIO

    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import (
        HRFlowable,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    dashboard = get_dashboard_gerencia(db, empresa_id)
    kpis = dashboard["kpis"]
    fecha_actual = datetime.now(timezone.utc).replace(tzinfo=None).strftime("%d/%m/%Y")
    W, H = letter

    # ── Colores corporativos ────────────────────────────────────────
    NAVY = colors.HexColor("#1B3A5C")
    BLUE = colors.HexColor("#2563EB")
    ACCENT = colors.HexColor("#0EA5E9")
    WHITE = colors.white
    GRAY_BG = colors.HexColor("#F8FAFC")
    GRAY_LN = colors.HexColor("#E2E8F0")
    MUTED = colors.HexColor("#64748B")

    # KPI colors
    C_GREEN = colors.HexColor("#DCFCE7")
    C_RED = colors.HexColor("#FEE2E2")
    C_YELLOW = colors.HexColor("#FEF9C3")
    C_BLUE = colors.HexColor("#DBEAFE")

    # ── Callbacks de página ─────────────────────────────────────────
    def header_footer(canvas_obj, doc_obj):
        canvas_obj.saveState()

        # Banda superior azul
        canvas_obj.setFillColor(NAVY)
        canvas_obj.rect(0, H - 72, W, 72, fill=True, stroke=False)

        # Línea de acento debajo de la banda
        canvas_obj.setFillColor(ACCENT)
        canvas_obj.rect(0, H - 76, W, 4, fill=True, stroke=False)

        # Texto PISST en la banda
        canvas_obj.setFillColor(WHITE)
        canvas_obj.setFont("Helvetica-Bold", 22)
        canvas_obj.drawString(0.75 * inch, H - 40, "PISST")

        canvas_obj.setFont("Helvetica", 9)
        canvas_obj.setFillColor(colors.HexColor("#A0C4E8"))
        canvas_obj.drawString(
            0.75 * inch,
            H - 56,
            "Plataforma Integral de Seguridad y Salud en el Trabajo",
        )

        # Fecha en la banda (derecha)
        canvas_obj.setFont("Helvetica", 9)
        canvas_obj.setFillColor(colors.HexColor("#CBD5E1"))
        canvas_obj.drawRightString(W - 0.75 * inch, H - 40, f"Generado: {fecha_actual}")
        canvas_obj.drawRightString(
            W - 0.75 * inch, H - 56, f"Período: {periodo.capitalize()}"
        )

        # Footer
        canvas_obj.setFillColor(GRAY_BG)
        canvas_obj.rect(0, 0, W, 36, fill=True, stroke=False)
        canvas_obj.setFillColor(GRAY_LN)
        canvas_obj.rect(0, 36, W, 1, fill=True, stroke=False)

        canvas_obj.setFont("Helvetica", 8)
        canvas_obj.setFillColor(MUTED)
        canvas_obj.drawString(
            0.75 * inch, 14, "PISST — Reporte generado automáticamente por el sistema."
        )
        canvas_obj.drawRightString(W - 0.75 * inch, 14, f"Página {doc_obj.page}")

        canvas_obj.restoreState()

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=0.75 * inch,
        leftMargin=0.75 * inch,
        topMargin=1.1 * inch,
        bottomMargin=0.7 * inch,
    )
    styles = getSampleStyleSheet()

    def s(nombre, size, color, bold=False, align=TA_LEFT, after=6, before=0):
        return ParagraphStyle(
            nombre,
            parent=styles["Normal"],
            fontSize=size,
            textColor=color,
            alignment=align,
            fontName="Helvetica-Bold" if bold else "Helvetica",
            spaceAfter=after,
            spaceBefore=before,
        )

    # ── Helpers de tabla ────────────────────────────────────────────
    def tabla_styled(data, col_widths, hdr_color):
        t = Table(data, colWidths=col_widths, repeatRows=1)
        style = [
            # Encabezado
            ("BACKGROUND", (0, 0), (-1, 0), hdr_color),
            ("TEXTCOLOR", (0, 0), (-1, 0), WHITE),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 10),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("TOPPADDING", (0, 0), (-1, 0), 9),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 9),
            # Filas
            ("FONTSIZE", (0, 1), (-1, -1), 9.5),
            ("TOPPADDING", (0, 1), (-1, -1), 7),
            ("BOTTOMPADDING", (0, 1), (-1, -1), 7),
            ("ALIGN", (1, 1), (-1, -1), "CENTER"),
            ("ALIGN", (0, 1), (0, -1), "LEFT"),
            ("LEFTPADDING", (0, 1), (0, -1), 10),
            ("GRID", (0, 0), (-1, -1), 0.4, GRAY_LN),
            ("LINEBELOW", (0, 0), (-1, 0), 1.5, hdr_color),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [WHITE, GRAY_BG]),
        ]
        t.setStyle(TableStyle(style))
        return t

    def kpi_cards(items):
        """items: lista de (label, value, bg_color)"""
        col_w = (W - 1.5 * inch) / len(items)
        row = []
        for label, value, bg in items:
            cell = Table(
                [
                    [
                        Paragraph(
                            str(value),
                            s("cv", 20, NAVY, bold=True, align=TA_CENTER, after=2),
                        )
                    ],
                    [Paragraph(label, s("cl", 8, MUTED, align=TA_CENTER, after=0))],
                ],
                colWidths=[col_w - 8],
            )
            cell.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, -1), bg),
                        ("TOPPADDING", (0, 0), (-1, -1), 10),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
                        ("LEFTPADDING", (0, 0), (-1, -1), 6),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                        ("ROUNDEDCORNERS", [4, 4, 4, 4]),
                        ("BOX", (0, 0), (-1, -1), 0.5, GRAY_LN),
                    ]
                )
            )
            row.append(cell)
        wrapper = Table([row], colWidths=[col_w] * len(items))
        wrapper.setStyle(
            TableStyle(
                [
                    ("LEFTPADDING", (0, 0), (-1, -1), 4),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                    ("TOPPADDING", (0, 0), (-1, -1), 0),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
                ]
            )
        )
        return wrapper

    # ── Contenido ───────────────────────────────────────────────────
    contenido = []

    # Título del reporte
    contenido.append(Spacer(1, 6))
    contenido.append(
        Paragraph(
            "Reporte Ejecutivo SG-SST",
            s("h1", 16, NAVY, bold=True, align=TA_CENTER, after=4),
        )
    )
    contenido.append(
        Paragraph(
            f"Período: <b>{periodo.capitalize()}</b>",
            s("h2", 10, MUTED, align=TA_CENTER, after=16),
        )
    )
    contenido.append(
        HRFlowable(width="100%", thickness=1.5, color=ACCENT, spaceAfter=20)
    )

    # ── Sección 1: Tarjetas KPI ─────────────────────────────────────
    contenido.append(
        Paragraph(
            "Indicadores Clave de Seguridad (KPIs)",
            s("sh", 11, NAVY, bold=True, after=10),
        )
    )

    contenido.append(
        kpi_cards(
            [
                ("Trabajadores", kpis["total_trabajadores"], C_BLUE),
                ("Accidentes (año)", kpis["total_accidentes"], C_RED),
                ("Días Perdidos", kpis["dias_perdidos"], C_YELLOW),
                ("Cumplimiento SG-SST", f"{dashboard['cumplimiento_sgsst']}%", C_GREEN),
            ]
        )
    )
    contenido.append(Spacer(1, 12))
    contenido.append(
        kpi_cards(
            [
                ("Incidentes Activos", dashboard["incidentes_activos"], C_RED),
                ("Incidentes Últ. Mes", dashboard["incidentes_ultimo_mes"], C_YELLOW),
                ("Capacitaciones", dashboard["total_capacitaciones"], C_BLUE),
                (
                    "Acciones Vencidas",
                    dashboard["acciones_vencidas"],
                    C_RED if dashboard["acciones_vencidas"] > 0 else C_GREEN,
                ),
            ]
        )
    )

    contenido.append(Spacer(1, 20))
    contenido.append(
        HRFlowable(width="100%", thickness=0.5, color=GRAY_LN, spaceAfter=16)
    )

    # ── Sección 2: Tabla KPIs detallados ───────────────────────────
    contenido.append(
        Paragraph(
            "Detalle de KPIs de Seguridad", s("sh", 11, NAVY, bold=True, after=10)
        )
    )

    data_kpis = [
        ["Indicador", "Valor", "Referencia"],
        ["Total Trabajadores Activos", str(kpis["total_trabajadores"]), "—"],
        ["Total Accidentes (año en curso)", str(kpis["total_accidentes"]), "Meta: 0"],
        ["Días Perdidos por Incapacidad", str(kpis["dias_perdidos"]), "Meta: < 30"],
        ["Tasa de Accidentalidad", f"{kpis['tasa_accidentalidad']}%", "Meta: < 5%"],
        ["Índice de Frecuencia (IF)", str(kpis["indice_frecuencia"]), "Meta: < 10"],
        ["Índice de Severidad (IS)", str(kpis["indice_severidad"]), "Meta: < 200"],
    ]
    cw = W - 1.5 * inch
    contenido.append(tabla_styled(data_kpis, [cw * 0.50, cw * 0.25, cw * 0.25], NAVY))
    contenido.append(Spacer(1, 20))

    # ── Sección 3: Resumen ejecutivo ────────────────────────────────
    contenido.append(
        HRFlowable(width="100%", thickness=0.5, color=GRAY_LN, spaceAfter=16)
    )
    contenido.append(
        Paragraph(
            "Resumen Ejecutivo del SG-SST", s("sh", 11, NAVY, bold=True, after=10)
        )
    )

    data_resumen = [
        ["Métrica", "Estado Actual", "Observación"],
        [
            "Cumplimiento SG-SST",
            f"{dashboard['cumplimiento_sgsst']}%",
            "Acciones correctivas completadas / total",
        ],
        [
            "Incidentes Activos",
            str(dashboard["incidentes_activos"]),
            "Sin estado cerrado",
        ],
        [
            "Incidentes Último Mes",
            str(dashboard["incidentes_ultimo_mes"]),
            "Período seleccionado",
        ],
        [
            "Capacitaciones Activas",
            str(dashboard["total_capacitaciones"]),
            "Programas vigentes",
        ],
        [
            "Acciones Correctivas Vencidas",
            str(dashboard["acciones_vencidas"]),
            "Requieren atención inmediata",
        ],
    ]
    contenido.append(
        tabla_styled(data_resumen, [cw * 0.38, cw * 0.22, cw * 0.40], BLUE)
    )
    contenido.append(Spacer(1, 10))

    # Nota legal
    contenido.append(
        HRFlowable(width="100%", thickness=0.5, color=GRAY_LN, spaceAfter=8)
    )
    contenido.append(
        Paragraph(
            "Este reporte fue generado automáticamente por la plataforma PISST con base en los datos registrados al momento de la descarga. "
            "Los indicadores se calculan según la metodología del Decreto 1072 de 2015 y la Resolución 0312 de 2019.",
            s("nota", 7.5, MUTED, align=TA_LEFT, after=0),
        )
    )

    doc.build(contenido, onFirstPage=header_footer, onLaterPages=header_footer)
    buffer.seek(0)
    return buffer


def generar_reporte_excel(db: Session, empresa_id: UUID, periodo: str):
    from io import BytesIO

    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    dashboard = get_dashboard_gerencia(db, empresa_id)
    kpis = dashboard["kpis"]
    fecha = datetime.now(timezone.utc).replace(tzinfo=None).strftime("%d/%m/%Y")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte PISST"

    # ── Paleta (misma que el PDF) ───────────────────────────────────
    NAVY = "1B3A5C"
    ACCENT = "0EA5E9"
    WHITE = "FFFFFF"
    GRAY_BG = "F8FAFC"
    GRAY_LN = "E2E8F0"
    MUTED = "64748B"
    TEXT = "1E293B"
    C_GREEN = "DCFCE7"
    C_RED = "FEE2E2"
    C_YELLOW = "FEF9C3"
    C_BLUE = "DBEAFE"
    C_GREEN_TXT = "166534"
    C_RED_TXT = "991B1B"
    C_YELLOW_TXT = "854D0E"
    C_BLUE_TXT = "1E40AF"

    # ── Estilos reutilizables ───────────────────────────────────────
    def font(bold=False, size=10, color=TEXT):
        return Font(name="Calibri", bold=bold, size=size, color=color)

    def fill(color):
        return PatternFill("solid", fgColor=color)

    def align(h="center", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    def border_thin(color=GRAY_LN):
        s = Side(style="thin", color=color)
        return Border(left=s, right=s, top=s, bottom=s)

    def border_bottom(color=ACCENT):
        return Border(bottom=Side(style="medium", color=color))

    def set_cell(
        ws,
        row,
        col,
        value,
        bold=False,
        size=10,
        color=TEXT,
        bg=None,
        h="center",
        v="center",
        wrap=False,
        brd=None,
    ):
        c = ws.cell(row=row, column=col, value=value)
        c.font = font(bold=bold, size=size, color=color)
        c.alignment = align(h=h, v=v, wrap=wrap)
        if bg:
            c.fill = fill(bg)
        if brd:
            c.border = brd
        return c

    # ── Anchos de columna ───────────────────────────────────────────
    ws.column_dimensions["A"].width = 4  # margen izq
    ws.column_dimensions["B"].width = 34  # etiqueta / métrica
    ws.column_dimensions["C"].width = 16  # valor
    ws.column_dimensions["D"].width = 24  # referencia / observación
    ws.column_dimensions["E"].width = 4  # margen der

    # ── Fila 1: banda de encabezado ─────────────────────────────────
    ws.row_dimensions[1].height = 14
    for col in range(1, 6):
        ws.cell(row=1, column=col).fill = fill(NAVY)

    # ── Fila 2: título PISST ────────────────────────────────────────
    ws.row_dimensions[2].height = 32
    ws.merge_cells("B2:D2")
    set_cell(
        ws,
        2,
        2,
        "PISST — Reporte Ejecutivo SG-SST",
        bold=True,
        size=16,
        color=WHITE,
        bg=NAVY,
        h="left",
    )
    for col in (1, 5):
        ws.cell(row=2, column=col).fill = fill(NAVY)

    # ── Fila 3: subtítulo ───────────────────────────────────────────
    ws.row_dimensions[3].height = 20
    ws.merge_cells("B3:D3")
    set_cell(
        ws,
        3,
        2,
        f"Período: {periodo.capitalize()}   |   Generado: {fecha}   |   Plataforma Integral de SST",
        size=9,
        color="A0C4E8",
        bg=NAVY,
        h="left",
    )
    for col in (1, 5):
        ws.cell(row=3, column=col).fill = fill(NAVY)

    # ── Fila 4: banda de acento ─────────────────────────────────────
    ws.row_dimensions[4].height = 5
    for col in range(1, 6):
        ws.cell(row=4, column=col).fill = fill(ACCENT)

    # ── Fila 5: espacio ─────────────────────────────────────────────
    ws.row_dimensions[5].height = 10

    # ── Helper: encabezado de sección ──────────────────────────────
    def seccion(titulo):
        r = ws.max_row + 1
        ws.row_dimensions[r].height = 22
        ws.merge_cells(f"B{r}:D{r}")
        set_cell(
            ws, r, 2, f"  {titulo}", bold=True, size=10, color=WHITE, bg=NAVY, h="left"
        )
        ws.cell(row=r, column=1).fill = fill(GRAY_BG)
        ws.cell(row=r, column=5).fill = fill(GRAY_BG)
        return r

    # ── Helper: fila de encabezado de tabla ─────────────────────────
    def tabla_header(cols):
        r = ws.max_row + 1
        ws.row_dimensions[r].height = 20
        for i, (col_idx, txt) in enumerate(cols):
            set_cell(
                ws,
                r,
                col_idx,
                txt,
                bold=True,
                size=9,
                color=WHITE,
                bg=NAVY,
                brd=border_thin("1B3A5C"),
            )
        ws.cell(row=r, column=1).fill = fill(GRAY_BG)
        ws.cell(row=r, column=5).fill = fill(GRAY_BG)
        return r

    # ── Helper: fila de dato ────────────────────────────────────────
    def fila_dato(label, valor, referencia=None, i=0, val_bg=None, val_color=TEXT):
        r = ws.max_row + 1
        ws.row_dimensions[r].height = 18
        bg = GRAY_BG if i % 2 == 0 else WHITE
        set_cell(
            ws,
            r,
            2,
            label,
            bold=False,
            size=9,
            color=TEXT,
            bg=bg,
            h="left",
            brd=border_thin(),
        )
        set_cell(
            ws,
            r,
            3,
            valor,
            bold=True,
            size=10,
            color=val_color,
            bg=val_bg or bg,
            brd=border_thin(),
        )
        if referencia is not None:
            set_cell(
                ws,
                r,
                4,
                referencia,
                size=8,
                color=MUTED,
                bg=bg,
                h="left",
                brd=border_thin(),
            )
        ws.cell(row=r, column=1).fill = fill(GRAY_BG)
        ws.cell(row=r, column=5).fill = fill(GRAY_BG)

    # ── Helper: tarjeta KPI (2 cols: etiqueta + valor coloreado) ────
    def tarjetas_kpi(items):
        r = ws.max_row + 1
        ws.row_dimensions[r].height = 14
        r2 = r + 1
        ws.row_dimensions[r2].height = 28
        r3 = r + 2
        ws.row_dimensions[r3].height = 16

        # items = [(label, valor, bg, txt_color), ...]
        # Distribuimos en B, C, D (3 columnas repartidas en 2 filas de tarjetas)
        cols_map = [2, 3, 4]
        for idx, (label, valor, bg, txt_c) in enumerate(items[:3]):
            col = cols_map[idx]
            # fila etiqueta
            set_cell(ws, r2, col, label, size=8, color=MUTED, bg=bg, h="center")
            # fila valor
            set_cell(
                ws,
                r3,
                col,
                str(valor),
                bold=True,
                size=13,
                color=txt_c,
                bg=bg,
                h="center",
            )
            ws.cell(row=r2, column=col).border = border_thin()
            ws.cell(row=r3, column=col).border = border_thin()

        for row in (r, r2, r3):
            ws.cell(row=row, column=1).fill = fill(GRAY_BG)
            ws.cell(row=row, column=5).fill = fill(GRAY_BG)

    # ── Espacio antes de sección 1 ──────────────────────────────────
    ws.append([])
    ws.row_dimensions[ws.max_row].height = 6

    # ── Sección 1: KPIs ─────────────────────────────────────────────
    seccion("INDICADORES CLAVE DE SEGURIDAD (KPIs)")
    ws.append([])
    ws.row_dimensions[ws.max_row].height = 6

    tabla_header([(2, "Indicador"), (3, "Valor"), (4, "Referencia / Meta")])

    kpi_rows = [
        (
            "Total Trabajadores Activos",
            kpis["total_trabajadores"],
            "—",
            C_BLUE,
            C_BLUE_TXT,
        ),
        (
            "Total Accidentes (año en curso)",
            kpis["total_accidentes"],
            "Meta: 0",
            C_RED,
            C_RED_TXT,
        ),
        (
            "Días Perdidos por Incapacidad",
            kpis["dias_perdidos"],
            "Meta: < 30",
            C_YELLOW,
            C_YELLOW_TXT,
        ),
        (
            "Tasa de Accidentalidad",
            f"{kpis['tasa_accidentalidad']}%",
            "Meta: < 5%",
            C_YELLOW,
            C_YELLOW_TXT,
        ),
        (
            "Índice de Frecuencia (IF)",
            str(kpis["indice_frecuencia"]),
            "Meta: < 10",
            C_RED,
            C_RED_TXT,
        ),
        (
            "Índice de Severidad (IS)",
            str(kpis["indice_severidad"]),
            "Meta: < 200",
            C_RED,
            C_RED_TXT,
        ),
    ]
    for i, (lbl, val, ref, bg, tc) in enumerate(kpi_rows):
        fila_dato(lbl, val, ref, i=i, val_bg=bg, val_color=tc)

    ws.append([])
    ws.row_dimensions[ws.max_row].height = 10

    # ── Sección 2: Resumen ejecutivo ────────────────────────────────
    seccion("RESUMEN EJECUTIVO DEL SG-SST")
    ws.append([])
    ws.row_dimensions[ws.max_row].height = 6

    tabla_header([(2, "Métrica"), (3, "Estado Actual"), (4, "Observación")])

    cumpl = dashboard["cumplimiento_sgsst"]
    cumpl_bg = C_GREEN if cumpl >= 80 else C_YELLOW if cumpl >= 50 else C_RED
    cumpl_tc = (
        C_GREEN_TXT if cumpl >= 80 else C_YELLOW_TXT if cumpl >= 50 else C_RED_TXT
    )

    venc = dashboard["acciones_vencidas"]
    venc_bg = C_RED if venc > 0 else C_GREEN
    venc_tc = C_RED_TXT if venc > 0 else C_GREEN_TXT

    resumen_rows = [
        (
            "Cumplimiento SG-SST",
            f"{cumpl}%",
            "Acciones completadas / total",
            cumpl_bg,
            cumpl_tc,
        ),
        (
            "Incidentes Activos",
            str(dashboard["incidentes_activos"]),
            "Sin estado cerrado",
            C_RED,
            C_RED_TXT,
        ),
        (
            "Incidentes Último Mes",
            str(dashboard["incidentes_ultimo_mes"]),
            "Período seleccionado",
            C_YELLOW,
            C_YELLOW_TXT,
        ),
        (
            "Capacitaciones Activas",
            str(dashboard["total_capacitaciones"]),
            "Programas vigentes",
            C_BLUE,
            C_BLUE_TXT,
        ),
        (
            "Acciones Correctivas Vencidas",
            str(venc),
            "Requieren atención inmediata",
            venc_bg,
            venc_tc,
        ),
    ]
    for i, (lbl, val, obs, bg, tc) in enumerate(resumen_rows):
        fila_dato(lbl, val, obs, i=i, val_bg=bg, val_color=tc)

    ws.append([])
    ws.row_dimensions[ws.max_row].height = 10

    # ── Footer ──────────────────────────────────────────────────────
    r_foot = ws.max_row + 1
    ws.row_dimensions[r_foot].height = 5
    for col in range(1, 6):
        ws.cell(row=r_foot, column=col).fill = fill(ACCENT)

    r_foot2 = r_foot + 1
    ws.row_dimensions[r_foot2].height = 18
    ws.merge_cells(f"B{r_foot2}:D{r_foot2}")
    set_cell(
        ws,
        r_foot2,
        2,
        f"PISST — Reporte generado automáticamente el {fecha}. "
        "Los indicadores se calculan según el Decreto 1072 de 2015 y la Resolución 0312 de 2019.",
        size=7.5,
        color=MUTED,
        bg=GRAY_BG,
        h="left",
        wrap=True,
    )
    ws.cell(row=r_foot2, column=1).fill = fill(GRAY_BG)
    ws.cell(row=r_foot2, column=5).fill = fill(GRAY_BG)

    # ── Ocultar líneas de cuadrícula ────────────────────────────────
    ws.sheet_view.showGridLines = False

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
